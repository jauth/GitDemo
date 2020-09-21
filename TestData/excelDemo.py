import openpyxl
# book = openpyxl.load_workbook("C:\\Users\\JD\\Documents\\PythonDemo.xlsx")  # Load excel sheet
# sheet = book.active  # set sheet to active sheet
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
#
# sheet.cell(row=2, column=2).value = "Jason Auth"
#
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row, "rows")
#
# print(sheet.max_column, "columns")
#
# print(sheet['A5'].value)
# print("--------------------------------------")
#
# for i in range(1, sheet.max_row+1):  # to get rows
#     if sheet.cell(row=i, column=1).value == "Testcase2":
#         for j in range(2, sheet.max_column+1):  # to get columns
#             print(sheet.cell(row=i, column=j).value)
#
#
# datadict = {}
# book = openpyxl.load_workbook("C:\\Users\\JD\\Documents\\PythonDemo.xlsx")  # Load excel sheet
# sheet = book.active  # set sheet to active sheet
#
# for i in range(1, sheet.max_row+1):  # to get rows
#     if sheet.cell(row=i, column=1) == "Testcase2":
#         for j in range(2, sheet.max_column+1):  # to get columns
#             # datadict["column header"] = cell value
#             datadict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
#             print([datadict])
