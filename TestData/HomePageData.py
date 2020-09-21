import openpyxl


class HomePageData:

    test_homepage_data = [({"name": "Jason Auth", "email": "jason.auth@gmail.com", "password": "Password1",
                            "gender": "Male"}),
                          ({"name": "Katherine Phillips", "email": "katiekat301@gmail.com", "password": "Password2",
                            "gender": "Female"}),
                          ({"name": "Troy Auth", "email": "troydauth@gmail.com", "password": "Password3",
                            "gender": "Male"})]

    @staticmethod  # static method so that the method can be called directly without creating an object
    def get_test_data(test_case_name):  # self not needed for static method

        datadict = {}
        book = openpyxl.load_workbook("C:\\Users\\JD\\Documents\\PythonDemo.xlsx")  # Load excel sheet
        sheet = book.active  # set sheet to active sheet

        for i in range(1, sheet.max_row+1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column+1):  # to get columns
                    # datadict["column header"] = cell value
                    datadict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    print([datadict])
        return [datadict]



